# -*- coding: utf-8 -*-
"""Offline builders for PDF/XLSX knowledge chunks and API embeddings."""

from __future__ import annotations

import json
import logging
import os
import re
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

import numpy as np
import pdfplumber
from openpyxl import load_workbook

from .api import OpenAICompatibleClient

logger = logging.getLogger(__name__)

QUESTION_TYPES = ("A1", "A2", "A3", "A4", "B1", "COMMON", "OTHER")
CONTENT_ROLES = (
    "type_definition",
    "type_requirement",
    "design_requirement",
    "general_principle",
    "example_good",
    "example_bad",
    "other",
)
CONTENT_ROLE_PRIORITY = (
    "example_bad",
    "example_good",
    "type_requirement",
    "type_definition",
    "design_requirement",
    "general_principle",
    "other",
)


def _normalize_text(text: str) -> str:
    cleaned = str(text or "").replace("\u3000", " ")
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _split_long_paragraph(text: str, max_chars: int = 500) -> List[str]:
    value = _normalize_text(text)
    if len(value) <= max_chars:
        return [value] if value else []
    pieces: List[str] = []
    start = 0
    while start < len(value):
        end = min(len(value), start + max_chars)
        window = value[start:end]
        cut = max(window.rfind(sep) for sep in ("。", "；", "\n", "，", " "))
        if cut < max_chars // 3:
            cut = len(window)
        piece = value[start : start + cut].strip()
        if piece:
            pieces.append(piece)
        start += cut
    return pieces


def split_pdf_page_into_chunks(page_text: str, min_chars: int = 50, max_chars: int = 500) -> List[str]:
    page_text = _normalize_text(page_text)
    if not page_text:
        return []
    raw_parts = [part.strip() for part in re.split(r"\n\s*\n", page_text) if part.strip()]
    if not raw_parts:
        raw_parts = [part.strip() for part in page_text.splitlines() if part.strip()]

    merged: List[str] = []
    buffer = ""
    for part in raw_parts:
        if len(part) > max_chars:
            if buffer:
                merged.append(buffer)
                buffer = ""
            merged.extend(_split_long_paragraph(part, max_chars=max_chars))
            continue
        if not buffer:
            buffer = part
            continue
        if len(buffer) < min_chars:
            candidate = f"{buffer}\n{part}".strip()
            if len(candidate) <= max_chars:
                buffer = candidate
                continue
            merged.append(buffer)
            buffer = part
            continue
        merged.append(buffer)
        buffer = part
    if buffer:
        merged.append(buffer)
    return [item for item in merged if item.strip()]


def _parse_pdf_tag_payload(raw: str, page: int = 0, chunk_id: str = "") -> Dict[str, object]:
    if not raw or not raw.strip():
        raise ValueError(f"[page={page}, chunk={chunk_id}] LLM 返回了空内容，无法解析 JSON。")
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"[page={page}, chunk={chunk_id}] LLM 返回了无效 JSON: {exc}") from exc
    qtypes = payload.get("question_type")
    role = payload.get("content_role")
    if not isinstance(qtypes, list) or not qtypes:
        raise ValueError(f"invalid question_type: {qtypes!r}")
    normalized_qtypes = []
    seen = set()
    for item in qtypes:
        value = str(item).upper().strip()
        if value not in QUESTION_TYPES:
            logger.warning("unknown question_type '%s', treating as OTHER", value)
            value = "OTHER"
        if value not in seen:
            seen.add(value)
            normalized_qtypes.append(value)
    role = payload.get("content_role")
    # 兼容 content_role 为字符串或列表（多选）
    roles_to_check = []
    if isinstance(role, list):
        roles_to_check = [str(r).strip() for r in role if r]
    elif isinstance(role, str):
        roles_to_check = [role.strip()]
    else:
        raise ValueError(f"invalid content_role format: {role!r}")

    normalized_roles = []
    seen_roles = set()
    for r in roles_to_check:
        if r not in CONTENT_ROLES:
            raise ValueError(f"unsupported content_role: {r}")
        if r not in seen_roles:
            seen_roles.add(r)
            normalized_roles.append(r)

    if not normalized_roles:
        raise ValueError("content_role cannot be empty")

    return {"question_type": normalized_qtypes, "content_role": normalized_roles}


def _tag_pdf_chunk_with_llm(
    client: OpenAICompatibleClient,
    model: str,
    text: str,
    page: int,
    chunk_index: int,
    previous_text: str = "",
    next_text: str = "",
) -> Dict[str, object]:
    user_prompt = f"""
请对下面这段 PDF 原文做标签标注，不允许改写原文。

页码: {page}
前一段（仅供上下文参考）:
{previous_text or "无"}

原文:
{text}

后一段（仅供上下文参考）:
{next_text or "无"}

输出要求:
1. 只输出 JSON。
2. question_type 只能从 {list(QUESTION_TYPES)} 中选择，可多选。
3. content_role 只能从 {list(CONTENT_ROLES)} 中选择，可多选。
4. 不允许改写原文，不允许创造新标签，不允许输出解释文字。

question_type 含义：
- A1/A2/A3/A4/B1：该段落明确属于对应题型
- COMMON：适用于多个题型或所有题型的通用命题规则
- OTHER：与 A1/A2/A3/A4/B1 单题评分关系弱

content_role 定义：
- type_definition：描述题型是什么、结构说明、组成方式
- type_requirement：描述某个题型的专属约束，必须/不能怎样
- design_requirement：描述题干、病例、设问、选项、答案、解析、干扰项等设计要求
- general_principle：跨题型通用质量原则
- example_good：正向例题、优秀样题、推荐示例
- example_bad：反例、问题样题、错误示例、带问题分析或修改建议的示例
- other：与单题 Judge 评分关系弱的其他内容

content_role 冲突优先级：
{ " > ".join(CONTENT_ROLE_PRIORITY) }

判断口诀：
- 像问题题、错误样题、带问题分析或修改建议 -> example_bad
- 像正常例题、正向样题 -> example_good
- 讲题型是什么 -> type_definition
- 讲某题型必须/不能怎样 -> type_requirement
- 讲怎么把题干、选项、病例、解析设计好 -> design_requirement
- 讲所有题型都适用的质量原则 -> general_principle
- 其他 -> other

JSON 结构:
{{
  "question_type": ["A2"],
  "content_role": "type_requirement"
}}
""".strip()
    chunk_id = f"p{page:03d}_c{chunk_index:02d}"

    # 对空响应 / JSON 解析错误做专门重试（网络层重试无法覆盖这种情况）
    last_error: Optional[Exception] = None
    for attempt in range(1, 6):
        try:
            raw = client.chat_complete_text(
                model=model,
                system_prompt="你是医学题目命题规范标注助手，只能输出 JSON。",
                user_prompt=user_prompt,
                temperature=0.0,
                max_tokens=256,
                response_format={"type": "json_object"},
            )
            return _parse_pdf_tag_payload(raw, page=page, chunk_id=chunk_id)
        except (json.JSONDecodeError, ValueError) as exc:
            last_error = exc
            if attempt < 3:
                wait = 2 ** (attempt - 1)
                logger.warning(
                    "[%s] tagging 第 %d 次重试（空/无效响应）: %s，%ds 后重试...",
                    chunk_id, attempt, exc, wait,
                )
                time.sleep(wait)
            # else: 最后一个 attempt 也失败了，抛出错误

    # 所有重试都失败，抛出带完整上下文的信息
    raise RuntimeError(
        f"[chunk={chunk_id}] tagging 5次重试均失败，最终错误: {last_error}"
    ) from last_error


def build_pdf_chunk_records(
    *,
    pdf_path: str,
    tagging_client: Optional[OpenAICompatibleClient] = None,
    tagging_model: str = "",
) -> List[Dict[str, object]]:
    if tagging_client is None or not str(tagging_model).strip():
        raise ValueError("PDF 标签构建必须使用 LLM 打标签；未提供可用 Judge API client 或 tagging_model。")

    records: List[Dict[str, object]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            text = _normalize_text(page.extract_text() or "")
            if not text:
                continue
            chunks = split_pdf_page_into_chunks(text)
            for chunk_index, chunk_text in enumerate(chunks, start=1):
                previous_text = chunks[chunk_index - 2] if chunk_index - 2 >= 0 else ""
                next_text = chunks[chunk_index] if chunk_index < len(chunks) else ""
                tags = _tag_pdf_chunk_with_llm(
                    tagging_client,
                    tagging_model,
                    chunk_text,
                    page_index,
                    chunk_index,
                    previous_text=previous_text,
                    next_text=next_text,
                )
                records.append(
                    {
                        "id": f"PDF_p{page_index:03d}_c{chunk_index:02d}",
                        "source": os.path.basename(pdf_path),
                        "page": page_index,
                        "text": chunk_text,
                        "question_type": tags["question_type"],
                        "content_role": tags["content_role"],
                    }
                )
    return records


def build_xlsx_chunk_records(xlsx_path: str) -> List[Dict[str, object]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    records: List[Dict[str, object]] = []
    for ws in workbook.worksheets:
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        for row_index, row in enumerate(rows, start=2):
            values = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row))) if headers[idx]}
            if not any(value is not None and str(value).strip() for value in values.values()):
                continue
            outline = [
                str(values.get(name, "")).strip()
                for name in ("一级大纲", "二级大纲", "三级大纲", "四级大纲", "五级大纲", "六级大纲")
                if str(values.get(name, "")).strip()
            ]
            text = " > ".join(
                item
                for item in [
                    str(values.get("专业名称", "")).strip(),
                    str(values.get("专业编码", "")).strip(),
                    *outline,
                ]
                if item
            )
            if not text:
                continue
            profession = str(values.get("专业名称", "")).strip()
            profession_code = str(values.get("专业编码", "")).strip()
            path_items = outline
            leaf = path_items[-1] if path_items else profession
            records.append(
                {
                    "id": f"XLSX_{ws.title}_{row_index}",
                    "source": os.path.basename(xlsx_path),
                    "sheet": ws.title,
                    "row": row_index,
                    "node_id": str(values.get("NODEID", "")).strip(),
                    "profession": profession,
                    "profession_code": profession_code,
                    "path": path_items,
                    "leaf": leaf,
                    "text": text,
                    "outline_path": text,
                    "NODEID": str(values.get("NODEID", "")).strip(),
                    "metadata": values,
                }
            )
    return records


def save_jsonl(records: Iterable[Dict[str, object]], output_path: str) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def build_embeddings_for_records(
    *,
    records: Sequence[Dict[str, object]],
    output_path: str,
    embedding_client: OpenAICompatibleClient,
    embedding_model: str,
    batch_size: int = 32,
) -> np.ndarray:
    if not embedding_model:
        raise ValueError("embedding_model is empty")
    texts = [str(record.get("text", "")) for record in records]
    vectors: List[List[float]] = []
    for start in range(0, len(texts), batch_size):
        batch = texts[start : start + batch_size]
        vectors.extend(embedding_client.embed_texts(model=embedding_model, texts=batch))
    matrix = np.asarray(vectors, dtype=np.float32)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    np.save(path, matrix)
    return matrix
